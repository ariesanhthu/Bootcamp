import os
import unittest
from openpyxl import load_workbook, Workbook

from exercises.day_1 import (
    ex3_0, ex3_1, ex3_2, ex3_3, ex3_4, ex3_5, ex3_6, ex3_7, ex3_8, ex3_9, ex3_10,
    ex35_1, ex35_2, ex35_3, ex35_4, ex35_5, ex35_6, ex35_7, ex35_8, ex35_9,
    ex4_1, ex4_2, ex4_3, ex4_4, ex4_5, ex4_6, ex4_7, ex4_8, ex4_9,
    ex5_1, ex5_2, ex5_3, ex5_4, ex5_5, ex5_6, ex5_7, ex5_8, ex5_9, ex5_10,
)
from exercises.day_2 import (
    ex6_1, ex6_2, ex6_3, ex6_4, ex6_5, ex6_6, ex6_7, ex6_8,  # ex6_9,
    ex69_1, ex69_2, ex69_3, ex69_4,
    ex7_1, ex7_2, ex7_3, ex7_4, ex7_5, ex7_6, ex7_7, ex7_8,
    ex8_3, ex8_4, ex8_5, ex8_7,
)


# Base Expercise
class TestExercise(unittest.TestCase):

    def _test_all(self, func, cases):
        for input_, expect in cases:
            output = func(input_)
            if output != expect:
                return False
        return True


# Test Exercise 3
class TestExercise3(TestExercise):

    def test_ex3_0(self):
        res = ex3_0.squared(5)
        if res == 25:
            return True
        return False

    def test_ex3_1(self):
        cases = [(1, 1), (5, 1), (9, 1), (24, 1000), (10, 10)]
        return self._test_all(self, ex3_1.solve, cases)

    def test_ex3_2(self):
        text = "P\nY\nM\nI"
        if ex3_2.solve(text) != "Pymi" or ex3_2.solve(ex3_2.data) != "Crossmyheart":
            return False

        return True

    def test_ex3_3(self):
        res = ex3_3.solve()
        if isinstance(res, list) is False or len(res) != len(range(1, 101)):
            return False

        cases = [(1, 1), (2, 2), (3, "Fizz"), (5, "Buzz"), (15, "FizzBuzz")]
        for num, value in cases:
            if res[num - 1] != value:
                return False

        return True

    def test_ex3_4(self):
        cases = [
            ("....slsslslsls...sls", "....slsslslsls.."),
            ("maria.data.mp9", "maria.data"),
            ("baitap.py", "baitap"),
        ]
        return self._test_all(self, ex3_4.solve, cases)

    def test_ex3_5(self):
        len_expected = len(ex3_5.data)
        res = ex3_5.solve(ex3_5.data)
        if isinstance(res, list) is False or len(res) != len_expected or res[0][0] != 1 or res[-1][0] != len_expected:
            return False

        return True

    def test_ex3_6(self):
        res = ex3_6.solve(2)
        if isinstance(res, tuple) is False or len(res) != 2:
            return False

        cases = [
            (1, ("January", 31)),
            (2, ("February", 28)),
            (3, ("March", 31)),
            (4, ("April", 30)),
            (7, ("July", 31)),
            (8, ("August", 31)),
            (9, ("September", 30)),
        ]
        return self._test_all(self, ex3_6.solve, cases)

    def test_ex3_7(self):
        len_expected = 19
        res = ex3_7.solve()
        if len(res) != len_expected or res[0] != "5 == 1 * 5" or res[-1] != "95 == 19 * 5":
            return False

        return True

    def test_ex3_8(self):
        cases = [
            ("ana", True),
            ("Civic", True),
            ("Python", False),
            ("", False),
            ("nN", False),
            ("P", False),
            (" P  ", False),
            (" P ", False),
            ("Able was I ere I saw Elba", True),
        ]
        return self._test_all(self, ex3_8.solve, cases)

    def test_ex3_9(self):
        len_expected = 23
        res = ex3_9.solve()
        if len(res) != len_expected or res[0] != [9, 1, 1] or res[-1] != [1, 9, 1]:  # Số bộ không đủ
            return False

        return True

    def test_ex3_10(self):
        cases = [
            (ex3_10.data, [0, 1]),
            (([3, 3], 6), [0, 1]),
            (([3, 2, 4], 6), [1, 2]),
        ]
        for test, result in cases:
            res = ex3_10.solve(*test)
            if res != result:
                return False

        return True


# Test Exercise 35
class TestExercise35(TestExercise):
    def test_ex35_1(self):
        result = ex35_1.solve(10)
        if len(result) != 10 or list(set(result))[0] != 2:
            return False
        return True

    def test_ex35_2(self):
        result = ex35_2.solve(5)
        if len(result) != 5:
            return False
        new_result = ex35_2.solve(5)
        # They are random, should be different
        if result == new_result or result[0] not in range(0, 10):
            return False

        # do the same with 10
        result = ex35_2.solve(10)
        if len(result) != 10:
            return False

        new_result = ex35_2.solve(10)
        if result == new_result:
            return False

        return True

    def test_ex35_3(self):
        result = ex35_3.solve(10)
        if result[:2] != [2, 4] or len(result) != 10:
            return False
        return True

    def test_ex35_4(self):
        result = ex35_4.solve(5)
        if len(result) != 5:
            return False
        new_result = ex35_4.solve(5)
        if result == new_result:
            return False

        return True

    def test_ex35_5(self):
        result = ex35_5.solve(12)
        if result[-1] != "111111111111" or result[0] != "111111":
            return False
        return True

    def test_ex35_6(self):
        result = ex35_6.solve(1000)
        if result != sum(int(i) for i in str(2 ** 1000)):
            return False
        return True

    def test_ex35_7(self):
        if ex35_7.solve(10) != 23 or ex35_7.solve(1000) != 233168:
            return False

        return True

    def test_ex35_8(self):
        result = ex35_8.solve(10)
        if result[:2] != "0*" or result[-2:] != "*9":
            return False

        N = 10
        matrix = [list(range(N)) for i in range(N)]
        for rowidx, row in enumerate(matrix):
            for idx, _ in enumerate(row):
                if idx == rowidx or idx == N - rowidx - 1:
                    row[idx] = str(rowidx)
                else:
                    row[idx] = "*"
        matrix = "\n".join(["".join(line) for line in matrix])

        if result != matrix:
            return False

        return True

    def test_ex35_9(self):
        res = ex35_9.solve(([], ""))
        if isinstance(res, list) is False or "__setitem__" not in res or "extend" in res:
            return False
        return True


# Test Exercise 4
class TestExercise4(TestExercise):
    def test_ex4_1(self):
        cases = [
            ("0.0.0.0", "00000000.00000000.00000000.00000000"),
            ("192.168.1.1", "11000000.10101000.00000001.00000001"),
            ("208.67.222.222", "11010000.01000011.11011110.11011110"),
        ]

        return self._test_all(self, ex4_1.solve, cases)

    def test_ex4_2(self):
        cases = [(420, '0o644'), (18, '0o22')]
        return self._test_all(self, ex4_2.solve, cases)

    def test_ex4_3(self):
        cases = [
            (
                ["knowledge", "hardwork", "discipline", "attitude"],
                [96, 98, 100, 100],
            )
        ]
        return self._test_all(self, ex4_3.solve, cases)

    def test_ex4_4(self):
        if ex4_4.solve() != 453542:
            return False

        return True

    def test_ex4_5(self):
        cases = [([1, 2, 3], (6, 6)), ([0, 1, 2], (3, 0))]

        return self._test_all(self, ex4_5.solve, cases)

    def test_ex4_6(self):
        ss3 = "10 Bé học lớp 3 lên lớp 4"
        ss = ", 60năm cuộc đời, 20 năm đầu, sung sướng0bao lâu"
        ss2 = "6năm0 cuộc đời, 20 năm đầu, sung sướng0bao lâu"

        import re

        pattern = re.compile(r"\d+")

        def _solve(ss):
            return [int(i) for i in pattern.findall(ss)]

        cases = [(i, _solve(i)) for i in [ss, ss2, ss3]]

        return self._test_all(ex4_6.solve, cases)

    def test_ex4_7(self):
        this, that, isgood = ex4_7.solve(1990, 1994)

        if "Canh Ngọ".lower() != this.lower() or "Giáp Tuất".lower() != that.lower() or ex4_7.GOOD != isgood:
            return False

        this, that, isgood = ex4_7.solve(1987, 1990)
        print(this, that, isgood)
        if "Đinh Mão".lower() != this or "Canh Ngọ".lower() != that or ex4_7.NEUTRAL != isgood:
            return False

        return True

    def test_ex4_8(self):
        result = ex4_8.solve()
        if (6, 8, 10) not in result or len(result) != 2:
            return False

        return True

    def test_ex4_9(self):
        cases = [([1, 3, 2], 3), ([42], 42)]
        return self._test_all(self, ex4_9.solve, cases)


# Test Exercise 5
class TestExercise5(TestExercise):
    def test_ex5_1(self):
        res = ex5_1.solve(ex5_1.data)
        if len(res) != len("Google") or res[0] != ("G", "#4885ed"):
            return False
        return True

    def test_ex5_2(self):
        hoang, duy, dai, tu = ex5_2.data
        print(hoang, duy, dai, tu)
        res = ex5_2.solve(ex5_2.data[::-1])
        if isinstance(res, list) is False or "languages" in duy:
            return False

        ns = {}
        for i in res:
            ns.update({i["name"]: i})

        newhoang, newduy, newdai, newtu = (
            ns["Hoang"],
            ns["Duy"],
            ns["Dai"],
            ns["Tu"],
        )
        if hoang["languages"] != newduy["languages"] or hoang["languages"] != newdai["languages"] \
                or hoang["languages"] != newtu["languages"] or newtu["girl_friend"] != "Do Anh" \
                or "Elixir" not in newhoang["languages"] or "girl_friend" in newduy:
            return False

        return True

    def test_ex5_3(self):
        res = ex5_3.solve(ex5_3.data)
        if len(res) != 10 or ("be", 5) not in res or ("can", 4) not in res:
            return False

        return True

    def test_ex5_4(self):
        expected = [
            "111111111111111111111111111111\n",
            "4\n",
            "111111111111111111111111111111\n",
            "8\n",
            "111111111111111111111111111111\n",
            "12\n",
            "111111111111111111111111111111\n",
            "16\n",
            "111111111111111111111111111111\n",
            "20\n",
            "111111111111111111111111111111\n",
            "59999984\n",
            "111111111111111111111111111111\n",
            "59999988\n",
            "111111111111111111111111111111\n",
            "59999992\n",
            "111111111111111111111111111111\n",
            "59999996\n",
            "111111111111111111111111111111\n",
            "60000000\n",
        ]
        import tempfile

        _, fn = tempfile.mkstemp()
        if ex5_4.solve(fn) != expected:
            return False

        return True

    def test_ex5_5(self):
        res = ex5_5.solve(ex5_5.datafile)
        msv, name, year, room = res[0]
        if msv >= ex5_5.MAGIC_NUMBER or year != 1990 or [("Dai", 5), ("Dung", 5), ("Duong", 5)] != [(e[1], e[-1]) for e
                                                                                                    in res[:3]]:
            return False

        return True

    def test_ex5_6(self):
        term1, term2 = ex5_6.data
        res = ex5_6.solve(term1, term2)
        if isinstance(res, dict) is False or res["python"] != 9 or res["math"] != 7 or res["data"] != 2:
            return False
        return True

    def test_ex5_7(self):
        cases = [
            ("a", "a"),
            ("aa", "aa2"),
            ("abbbccccdddd", "abb3cc4dd4"),
            ("xxyyyxyyx", "xx2yy3xyy2x"),
        ]
        return self._test_all(self, ex5_7.solve, cases)

    def test_ex5_8(self):
        ascii_, _, tabcp, newlinecp, spacecp = ex5_8.solve()
        if ascii_[:3] != [(33, "!"), (34, '"'), (35, "#")] or ascii_[-4:] != [(49, "1"), (50, "2"), (51, "3"),
                                                                              (52, "4")] \
                or tabcp != ord("\t") or spacecp != ord(" ") or newlinecp != ord("\n"):
            return False

        return True

    def test_ex5_9(self):
        start_with_h, more_than_1mil = ex5_9.solve(ex5_9.data)
        if start_with_h[-2:] != [("Hải Phòng", 1904100), ("Hậu Giang", 769700)] \
                or more_than_1mil[:2] != [("TP. Hồ Chí Minh", 7681700), ("Hà Nội", 6844100)]:
            return False
        return True

    def test_ex5_10(self):
        # TODO: code

        return False


# Test Exercise 6
class TestExercise6(TestExercise):
    def test_ex6_1(self):
        expected = [
            ("Nam", "1,230"),
            ("Pikalong", "35,670"),
            ("Phan Quan", "2,165,110"),
            ("Maria", "90,570"),
            ("Trump", "138,000"),
        ]
        len_expected = len(expected)
        res = ex6_1.solve(
            {
                "usages": [
                    ("nam", "1"),
                    ("pikalong", "29"),
                    ("phan quan", "1235"),
                    ("maria", "69"),
                    ("trump", "100"),
                ],
                "prices": ex6_1.data,
            }
        )
        if isinstance(res, list) is False or len(res) != len_expected or res != expected:
            return False

        return True

    def test_ex6_2(self):
        data = ["meo", "bo", "ga", "cho", "chim", "gau", "voi"]
        expected = list(zip(*[iter(data)] * 2))
        res = ex6_2.solve(data, 2)
        if isinstance(res, list) is False or isinstance(res[0], tuple) or res != expected \
                or ex6_2.solve(data, 3) != list(zip(*[iter(data)] * 3)):
            return False
        return True

    def test_ex6_3(self):
        expected = ("2017-05-25", 76454277.83)
        if ex6_3.solve() != expected:
            return False
        return True

    def test_ex6_4(self):
        res = ex6_4.solve()
        expected = sum(range(1, 7)) + 0.5
        if res != expected:
            return False

        return True

    def test_ex6_5(self):
        res = ex6_5.solve(ex6_5.data)
        if isinstance(res, list) is False or isinstance(res[0], dict) is False \
                or "hvnsweeting" not in [n["login"] for n in res] \
                or "https://github.com/thedrow" not in [n["html_url"] for n in res] \
                or "https://github.com/hvnsweeting" not in [n["html_url"] for n in res]:
            return False

        return True

    def test_ex6_6(self):
        res = ex6_6.solve("pallets")
        if "flask" not in res:
            return False

        return True

    def test_ex6_7(self):
        prefix = (
            "       1       1     0o1     0x1\n"
            "       2      10     0o2     0x2\n"
        )
        suffix = (
            "      18   10010    0o22    0x12\n"
            "      19   10011    0o23    0x13\n"
        )
        res = ex6_7.solve(range(1, 20))

        if prefix != res[: len(prefix)] or suffix != res[-len(suffix):]:
            return False

        return False

    def test_ex6_8(self):
        # TODO: code
        return False

    def test_ex6_9(self):
        # TODO: code
        return False


# Test Exercise 69
class TestExercise69(TestExercise):
    def test_ex69_1(self):
        res = ex69_1.solve(range(10))
        if isinstance(res, list) is False or res != [0, 2, 4, 6, 8, 10, 12, 14, 16, 18]:
            return False
        return True

    def test_ex69_2(self):
        res = ex69_2.solve(range(10))
        if isinstance(res, list) is False or res != [1, 3, 5, 7, 9]:
            return False

        return True

    def test_ex69_3(self):
        res = ex69_3.solve(range(1, 6))
        if isinstance(res, int) is False or res != 120:
            return False
        return True

    def test_ex69_4(self):
        # TODO: code
        return False


# Test Exercise 7
class TestExercise7(TestExercise):
    def test_ex7_1(self):
        if ex7_1.solve("1/3", "6/9") != 1.0 or ex7_1.solve("1/10", "1/10", "1/10") != 0.3:
            return False

        return True

    def test_ex7_2(self):
        name, HP = ex7_2.solve(
            ex7_2.Fighter("PyMi", 100), ex7_2.Fighter("Foo", 100)
        )
        if name not in ("PyMi", "Foo") or HP <= 0:
            return False

    def test_ex7_3(self):
        import json
        import os
        import pickle
        import yaml

        res = ex7_3.solve()
        len_expected = 4
        if res != len_expected:
            return False

        cwd = os.getcwd()
        current_dir = os.path.basename(cwd)
        if current_dir == "pyfml":
            fmt_path = "exercises/{}"
        elif current_dir == "test":
            fmt_path = "exercises/{}"
        else:
            raise Exception("Please run test under pyfml directory")
        cases = ("event.json", "event.yaml", "event.pkl")
        for case in cases:
            print(fmt_path)
            if fmt_path.format(case) is not True:
                return False

        with open(fmt_path.format("event.json")) as f:
            data = json.load(f)
        with open(fmt_path.format("event.yaml")) as f:
            out_yaml = yaml.safe_load(f)
            if data != out_yaml:
                return False

        with open(fmt_path.format("event.pkl"), "rb") as f:
            out_pkl = pickle.load(f)
            if data != out_pkl:
                return False

        return True

    def test_ex7_4(self):
        import hashlib

        r = hashlib.md5(ex7_4.solve().encode()).hexdigest()

        if r != "9fd583bca43a756b86dd74706afb24d8":
            return False

        return True

    def test_ex7_5(self):
        import os

        result = ex7_5.solve()
        if os.__file__ != result[0]:
            return False
        return True

    def test_ex7_6(self):
        import string

        res = ex7_6.solve(12)
        if len(res) != 12 or all(
                [
                    any(map(lambda s: s.isdigit(), res)),
                    any(map(lambda s: s.islower(), res)),
                    any(map(lambda s: s.isupper(), res)),
                    any(map(lambda s: s in string.punctuation, res)),
                ]
        ) is False or res == ex7_6.solve(12):
            return False
        return True

    def test_ex7_7(self):
        res = ex7_7.solve(ex7_7.data)
        if isinstance(res, list) is False or isinstance(res[0], tuple) is False or ("Singapore", 4) not in res:
            return False

        return True

    def test_ex7_8(self):
        # TODO code
        return False


# Test Exercise 8
class TestExercise8(TestExercise):
    def test_ex8_1(self):
        return False
        # N = 3
        # times, total_time = ex8_1.solve(N)
        # if len(times) != N or N > total_time:
        #     return False
        #
        # return True

    def test_ex8_2(self):
        return False
        # conf = configparser.ConfigParser()
        # conf.read(os.path.join(os.path.dirname(__file__), "../setup.cfg"))
        # n = conf.getint("flake8", "max-line-length")
        #
        # res = ex8_2.solve("-h", __file__)
        # if n != len(res):
        #     return False
        #
        # with open(__file__) as f:
        #     line = f.readline()
        #
        # if line.strip() != res[0].strip():
        #     return False
        #
        # _, fn = tempfile.mkstemp()
        # with open(fn, "w") as f:
        #     for i in range(100):
        #         f.write(str(i) + "\n")
        # last_n_lines = []
        # with open(fn) as f:
        #     for line in f:
        #         last_n_lines.append(line)
        #         if len(last_n_lines) > n:
        #             last_n_lines.pop(0)
        # res = ex8_2.solve("-t", fn)
        # if last_n_lines != res:
        #     return False
        #
        # return True

    def test_ex8_3(self):
        res = ex8_3.solve(ex8_3.data)
        if len(res) != len(ex8_3.data) or res[0] != ex8_3.data[0].upper():
            return False

        return True

    def test_ex8_4(self):
        timer_res = ex8_4.solve()
        if timer_res <= 1.0:
            return False
        return True

    def test_ex8_5(self):
        test_res = ex8_5.solve()
        if test_res is not True:
            return False

        return True

    def test_ex8_7(self):
        # TODO: code
        return False

    def test_ex8_8(self):
        return False
        # cases = [
        #     ("02/03/16", "0.3.1"),
        #     ("09/06/16", "8.2.0"),
        #     ("06/23/17", "18.3.3"),
        # ]
        #
        # return self._test_all(ex8_8.solve, cases)

    def __total_line_suffix(self, path):
        """
        :param path: path folder
        :rtype dict:
        """
        file_suffix = {}
        for root, _, files in os.walk(path):
            for file in files:
                path_file = os.path.join(root, file)
                _suffix = path_file[path_file.rfind("."):]
                if not _suffix:
                    continue
                if _suffix not in file_suffix:
                    file_suffix.update({_suffix: 0})

                try:
                    with open(file) as f:
                        file_suffix[_suffix] += sum(1 for _ in f)
                except IOError:
                    continue

        return file_suffix

    def test_ex8_9(self):
        return False
        # expected = self.__total_line_suffix(ex8_9.PATH)
        # res = ex8_9.solve(ex8_9.PATH)
        # if expected != res:
        #     return False
        #
        # return True

    def test_ex8_10(self):
        # TODO: code
        return False


# TITLE FILE REPORT
STT = "A"
NAME = "B"

EX3 = "C"  # Exercises 3
EX3_SUCCESS = "D"  # Exercises 3 success
EX3_ERROR = "E"  # Exercises 3 error

EX35 = "F"  # Exercises 35
EX35_SUCCESS = "G"  # Exercises 35 success
EX35_ERROR = "H"  # Exercises 35 error

EX4 = "I"  # Exercises 4
EX4_SUCCESS = "J"  # Exercises 4 success
EX4_ERROR = "K"  # Exercises 4 error

EX5 = "L"  # Exercises 5
EX5_SUCCESS = "M"  # Exercises 5 success
EX5_ERROR = "N"  # Exercises 5 error

EX6 = "O"  # Exercises 6
EX6_SUCCESS = "P"  # Exercises 6 success
EX6_ERROR = "Q"  # Exercises 6 error

EX69 = "R"  # Exercises 69
EX69_SUCCESS = "S"  # Exercises 69 success
EX69_ERROR = "T"  # Exercises 69 error

EX7 = "U"  # Exercises 7
EX7_SUCCESS = "V"  # Exercises 7 success
EX7_ERROR = "W"  # Exercises 7 error

EX8 = "X"  # Exercises 8
EX8_SUCCESS = "Y"  # Exercises 8 success
EX8_ERROR = "Z"  # Exercises 8 error


def set_data_to_excel(ws_file, idx, key_exercise, exercise_obj):
    data_exercise = data.get(key_exercise)
    total = exercise_obj.get("total")
    success = exercise_obj.get("success")
    error = exercise_obj.get("error")
    ws_file[f"{total}{idx}"] = data_exercise.get('total')
    ws_file[f"{success}{idx}"] = data_exercise.get('success')
    ws_file[f"{error}{idx}"] = data_exercise.get('error')


if __name__ == "__main__":
    all_class = [TestExercise3, TestExercise35, TestExercise4, TestExercise5, TestExercise6, TestExercise69,
                 TestExercise7, TestExercise8]

    data = dict()
    for name_class in all_class:
        all_func = name_class.__dict__
        key = name_class.__name__.lower()
        data.update({
            key: {
                "total": 0,
                "success": 0,
                "error": 0,
                "func_error_list": []
            }
        })

        obj_class = name_class
        for name_fu, name_fuc in all_func.items():
            if "__" not in name_fu:
                try:
                    data[key]['total'] += 1
                    success_flag = name_fuc(obj_class)
                    if success_flag is not True:
                        data[key]['error'] += 1
                        data[key]['func_error_list'].append(name_fu)

                except:
                    data[key]['error'] += 1
                    data[key]['func_error_list'].append(name_fu)

        data[key]['success'] = data[key]['total'] - data[key]['error']

    print('Enter your name:')
    name = input()
    if name != "":
        wb = load_workbook("report_ex.xlsx")
        ws = wb["Sheet1"]

        row_count = ws.max_row
        index = row_count + 1
        stt = row_count - 1

        # đổ data vào file excel
        ws[f"{STT}{index}"] = stt
        ws[f"{NAME}{index}"] = name

        # exercise
        exercise_list = [
            {
                "name": "testexercise3",
                "key_exercise": {
                    "total": EX3,
                    "success": EX3_SUCCESS,
                    "error": EX3_ERROR
                }
            },
            {
                "name": "testexercise35",
                "key_exercise": {
                    "total": EX35,
                    "success": EX35_SUCCESS,
                    "error": EX35_ERROR
                }
            },
            {
                "name": "testexercise4",
                "key_exercise": {
                    "total": EX4,
                    "success": EX4_SUCCESS,
                    "error": EX4_ERROR
                }
            },
            {
                "name": "testexercise5",
                "key_exercise": {
                    "total": EX5,
                    "success": EX5_SUCCESS,
                    "error": EX5_ERROR
                }
            },
            {
                "name": "testexercise6",
                "key_exercise": {
                    "total": EX6,
                    "success": EX6_SUCCESS,
                    "error": EX6_ERROR
                }
            },
            {
                "name": "testexercise69",
                "key_exercise": {
                    "total": EX69,
                    "success": EX69_SUCCESS,
                    "error": EX69_ERROR
                }
            },
            {
                "name": "testexercise7",
                "key_exercise": {
                    "total": EX7,
                    "success": EX7_SUCCESS,
                    "error": EX7_ERROR
                }
            },
            {
                "name": "testexercise8",
                "key_exercise": {
                    "total": EX8,
                    "success": EX8_SUCCESS,
                    "error": EX8_ERROR
                }
            }
        ]
        for exercise in exercise_list:
            set_data_to_excel(ws, index, exercise['name'], exercise['key_exercise'])

        wb.save("report_ex.xlsx")

    print("Save data success !!!!!")
